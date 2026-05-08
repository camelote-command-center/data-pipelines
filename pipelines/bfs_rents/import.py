#!/usr/bin/env python3
"""
BFS Average Rents — Import Pipeline

Pulls four BFS DAM Strukturerhebung (SE) rent xlsx assets and upserts into two
bronze tables on re-LLM:

    bronze_ch.bfs_average_rents          (canton × rooms × duration_class)
    bronze_ch.bfs_rents_by_owner_type    (canton × owner_type, 3-yr cumulated)

Sources:
  - 36398434  (FR) "Loyer moyen en francs selon le nombre de pièces, par canton"
                — yearly sheets 2000, 2003, 2010-2024, all 26 cantons + Suisse
                → bfs_average_rents (avg_rent_chf, duration_class='_all')

  - 36398449  (FR) "Loyer moyen par m2 en francs selon le nombre de pièces, par canton"
                — yearly sheets 2012-2024, same canton×rooms grain
                → bfs_average_rents (rent_per_m2_chf, duration_class='_all')

  - 36398531  (FR) "Loyer moyen selon la durée de la location et le nombre de pièces"
                — yearly sheets 2020-2024, 6 duration classes, no canton dim
                → bfs_average_rents (avg_rent_chf, canton_code='CH', duration_class set)

  - 36398506  (FR) "Type de propriétaire des logements de locataires, par canton, 2023-2025 cumulé"
                — single sheet, 6 owner types, all cantons + Suisse
                → bfs_rents_by_owner_type

The legacy DE asset 24129085 is superseded by 36398434 (covers same 2000-2021
range plus 2022/2023/2024). Same conflict key, so re-running both is safe.

DATA SAFETY:
    - UPSERT only. Never truncates or deletes.
    - Idempotent via the (year,canton,rooms,duration_class) and
      (period_label,canton_code,owner_type) unique constraints.

Environment variables:
    RE_LLM_SUPABASE_URL              - re-LLM Supabase project URL (required)
    RE_LLM_SUPABASE_SERVICE_ROLE_KEY - service_role key (required)
    RE_LLM_SCHEMA                    - target schema (default: bronze_ch)
    CAMELOTE_SUPABASE_URL            - Command center URL (optional, for metadata)
    CAMELOTE_SUPABASE_KEY            - Command center key (optional)
"""

from __future__ import annotations

import io
import os
import sys
import time

import openpyxl
import requests

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
from shared.supabase_client import batch_upsert  # noqa: E402

try:
    from shared.freshness import update_dataset_meta  # noqa: E402
except Exception:  # optional
    update_dataset_meta = None


# ──────────────────────────────────────────────────────────────
# Source constants
# ──────────────────────────────────────────────────────────────

DAM_MASTER = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/{asset_id}/master"

ASSET_CANTON_ROOMS_CHF       = 36398434  # FR, 2000-2024 — avg rent by canton×rooms
ASSET_CANTON_ROOMS_PER_M2    = 36398449  # FR, 2012-2024 — rent/m² by canton×rooms
ASSET_DURATION_ROOMS         = 36398531  # FR, 2020-2024 — avg rent by rooms×duration
ASSET_OWNER_TYPE_BY_CANTON   = 36398506  # FR, 2023-2025 cumulé — owner-type share by canton

TABLE_AVG    = "bfs_average_rents"
TABLE_OWNER  = "bfs_rents_by_owner_type"

CONFLICT_AVG   = "year,canton_code,rooms,duration_class"
CONFLICT_OWNER = "period_label,canton_code,owner_type"

DURATION_ALL_SENTINEL = "_all"

# Canton-name (FR) → 2-letter ISO code
CANTON_FR = {
    "Suisse": "CH",
    "Zurich": "ZH",
    "Berne": "BE",
    "Lucerne": "LU",
    "Uri": "UR",
    "Schwyz": "SZ",
    "Obwald": "OW",
    "Nidwald": "NW",
    "Glaris": "GL",
    "Zoug": "ZG",
    "Fribourg": "FR",
    "Soleure": "SO",
    "Bâle-Ville": "BS",
    "Bâle-Campagne": "BL",
    "Schaffhouse": "SH",
    "Appenzell Rh.-Ext.": "AR",
    "Appenzell Rh.-Int.": "AI",
    "Saint-Gall": "SG",
    "Grisons": "GR",
    "Argovie": "AG",
    "Thurgovie": "TG",
    "Tessin": "TI",
    "Vaud": "VD",
    "Valais": "VS",
    "Neuchâtel": "NE",
    "Genève": "GE",
    "Jura": "JU",
}

# Duration label normalisation. Source FR labels → snake_case canonical.
# Lookup is on the *cleaned* label (footnote markers like '3)' stripped).
DURATION_FR_NORMAL = {
    "Total": "total",
    "Moins de 2 ans dans des logements neufs": "lt_2y_new_dwellings",
    "Moins de 2 ans dans des logements construits il y a 2 ans ou plus":
        "lt_2y_existing_dwellings",
    "2 à 5 ans": "2_5y",
    "6-10 ans": "6_10y",
    "11-20 ans": "11_20y",
    "21 ans et plus": "21y_plus",
}


def _clean_label(label: str) -> str:
    """Strip BFS footnote markers and trailing whitespace.

    BFS labels often carry footnote refs like 'Moins de 2 ans dans des logements neufs3)'.
    Returns the label with trailing digit-paren footnotes removed.
    """
    import re as _re
    s = str(label).strip()
    # Remove trailing footnote markers like '3)', '12)', '*'
    s = _re.sub(r"\s*\d+\)\s*$", "", s)
    s = _re.sub(r"\s*\*+\s*$", "", s)
    return s.strip()


# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────

def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("…", "X", "*", "()") or s.startswith("(") or s.startswith("X"):
            return None
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _download_xlsx(asset_id: int) -> bytes:
    url = DAM_MASTER.format(asset_id=asset_id)
    print(f"  Downloading asset {asset_id} …")
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=120)
    r.raise_for_status()
    print(f"    {len(r.content):,} bytes")
    return r.content


def _detect_room_columns(header_row3) -> list[tuple[int, int]]:
    """Inspect row 3 of a canton-table sheet and yield (col_idx, room_count) pairs.

    Header layout: row 3 has room labels at every other column starting at col 1.
        cols: 1=Total, 3=1, 5=2, 7=3, 9=4, 11=5, 13=6+, 15=7+ (varies per asset)
    Each label cell is followed by an empty cell (the confidence-interval header
    sits on row 4 below). Convention: rooms=0 means "Total".
    """
    pairs: list[tuple[int, int]] = []
    for col_idx in range(1, len(header_row3)):
        label = header_row3[col_idx]
        if label is None or label == "":
            continue
        text = str(label).strip()
        if text.lower() == "total":
            pairs.append((col_idx, 0))
        else:
            # "1", "2", "6+", "7+", "8 et plus"
            digits = "".join(ch for ch in text if ch.isdigit())
            if digits:
                pairs.append((col_idx, int(digits)))
    return pairs


# ──────────────────────────────────────────────────────────────
# Parsing — canton × rooms (CHF or CHF/m²)
# ──────────────────────────────────────────────────────────────

def parse_canton_rooms(xlsx_bytes: bytes, value_field: str, source_asset_id: int) -> list[dict]:
    """Parse a canton × rooms xlsx (per-year sheets) into long format.

    `value_field`: 'average_rent_chf' or 'rent_per_m2_chf'.
    `source_asset_id`: stamped into source_file for provenance.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    out: list[dict] = []

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue
        room_cols = _detect_room_columns(rows[3])
        if not room_cols:
            continue
        for row in rows[5:]:
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()
            canton_code = CANTON_FR.get(label)
            if not canton_code:
                continue  # footer / annotation row
            for col_idx, rooms in room_cols:
                if col_idx >= len(row):
                    continue
                v = _to_float(row[col_idx])
                if v is None:
                    continue
                out.append({
                    "year": year,
                    "canton_code": canton_code,
                    "rooms": rooms,
                    "duration_class": DURATION_ALL_SENTINEL,
                    value_field: v,
                    "source_file": f"bfs_dam_asset_{source_asset_id}_{sheet_name}",
                })
    return out


# ──────────────────────────────────────────────────────────────
# Parsing — rooms × duration (no canton)
# ──────────────────────────────────────────────────────────────

def parse_rooms_duration(xlsx_bytes: bytes, source_asset_id: int) -> list[dict]:
    """Parse a rooms × duration xlsx (per-year sheets) into long format.

    All rows pin to canton_code='CH' since this asset has no canton dim.
    duration_class is the snake_case normalized label.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    out: list[dict] = []

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue
        room_cols = _detect_room_columns(rows[3])
        if not room_cols:
            continue
        for row in rows[5:]:
            if not row or not row[0]:
                continue
            label = _clean_label(row[0])
            duration = DURATION_FR_NORMAL.get(label)
            if duration is None:
                # Footer / annotation row
                continue
            for col_idx, rooms in room_cols:
                if col_idx >= len(row):
                    continue
                v = _to_float(row[col_idx])
                if v is None:
                    continue
                out.append({
                    "year": year,
                    "canton_code": "CH",
                    "rooms": rooms,
                    "duration_class": duration,
                    "average_rent_chf": v,
                    "source_file": f"bfs_dam_asset_{source_asset_id}_{sheet_name}",
                })
    return out


# ──────────────────────────────────────────────────────────────
# Parsing — owner type by canton (single 3-yr cumulative sheet)
# ──────────────────────────────────────────────────────────────

def parse_owner_type(xlsx_bytes: bytes, source_asset_id: int) -> list[dict]:
    """Parse the 'Type de propriétaire' xlsx into the owner-type sibling table."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    if not wb.sheetnames:
        return []
    sheet_name = wb.sheetnames[0]  # e.g. '2023-2025'
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return []

    # Owner-type labels live on row 2, every other column starting at col 1
    owner_cols: list[tuple[int, str]] = []
    for col_idx in range(1, len(rows[2])):
        v = rows[2][col_idx]
        if v and str(v).strip():
            owner_cols.append((col_idx, str(v).strip()))

    if not owner_cols:
        return []

    out: list[dict] = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        canton_code = CANTON_FR.get(label)
        if not canton_code:
            continue
        for col_idx, owner_type in owner_cols:
            if col_idx >= len(row):
                continue
            share = _to_float(row[col_idx])
            ci = _to_float(row[col_idx + 1]) if col_idx + 1 < len(row) else None
            if share is None:
                continue
            out.append({
                "period_label": sheet_name,  # e.g. '2023-2025'
                "canton_code": canton_code,
                "owner_type": owner_type,
                "share_pct": share,
                "confidence_interval": ci,
                "source_asset_id": source_asset_id,
            })
    return out


# ──────────────────────────────────────────────────────────────
# Merge canton×rooms CHF + CHF/m² records into single rows
# ──────────────────────────────────────────────────────────────

def merge_canton_rooms(
    chf_records: list[dict], per_m2_records: list[dict]
) -> list[dict]:
    """Two parses produce records on the same (year, canton, rooms, duration_class)
    grain — one with average_rent_chf, the other with rent_per_m2_chf. Merge on key.
    """
    by_key: dict[tuple, dict] = {}
    for rec in chf_records + per_m2_records:
        k = (rec["year"], rec["canton_code"], rec["rooms"], rec["duration_class"])
        if k not in by_key:
            by_key[k] = {
                "year": rec["year"],
                "canton_code": rec["canton_code"],
                "rooms": rec["rooms"],
                "duration_class": rec["duration_class"],
                "average_rent_chf": None,
                "rent_per_m2_chf": None,
                "source_file": rec.get("source_file"),
            }
        if "average_rent_chf" in rec and rec["average_rent_chf"] is not None:
            by_key[k]["average_rent_chf"] = rec["average_rent_chf"]
        if "rent_per_m2_chf" in rec and rec["rent_per_m2_chf"] is not None:
            by_key[k]["rent_per_m2_chf"] = rec["rent_per_m2_chf"]
    return list(by_key.values())


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def _required(name: str) -> str:
    v = os.environ.get(name, "")
    if not v:
        print(f"ERROR: env var {name} required")
        sys.exit(1)
    return v


def main():
    url = _required("RE_LLM_SUPABASE_URL")
    key = _required("RE_LLM_SUPABASE_SERVICE_ROLE_KEY")
    schema = os.environ.get("RE_LLM_SCHEMA", "bronze_ch")

    print("=" * 60)
    print("  BFS Rents Pipeline")
    print(f"  Target: {schema}.{TABLE_AVG} + {schema}.{TABLE_OWNER}")
    print("=" * 60)

    t0 = time.time()

    # ── Sources 1 & 2: canton×rooms CHF and CHF/m² → bfs_average_rents ──
    print("\n[1] canton×rooms CHF (asset {})".format(ASSET_CANTON_ROOMS_CHF))
    chf_bytes = _download_xlsx(ASSET_CANTON_ROOMS_CHF)
    chf_recs = parse_canton_rooms(chf_bytes, "average_rent_chf", ASSET_CANTON_ROOMS_CHF)
    print(f"  parsed: {len(chf_recs):,}")

    print("\n[2] canton×rooms CHF/m² (asset {})".format(ASSET_CANTON_ROOMS_PER_M2))
    m2_bytes = _download_xlsx(ASSET_CANTON_ROOMS_PER_M2)
    m2_recs = parse_canton_rooms(m2_bytes, "rent_per_m2_chf", ASSET_CANTON_ROOMS_PER_M2)
    print(f"  parsed: {len(m2_recs):,}")

    canton_rooms = merge_canton_rooms(chf_recs, m2_recs)
    print(f"  merged into {len(canton_rooms):,} rows")

    # ── Source 3: rooms×duration → bfs_average_rents (canton_code='CH') ──
    print("\n[3] rooms×duration (asset {})".format(ASSET_DURATION_ROOMS))
    dur_bytes = _download_xlsx(ASSET_DURATION_ROOMS)
    dur_recs = parse_rooms_duration(dur_bytes, ASSET_DURATION_ROOMS)
    print(f"  parsed: {len(dur_recs):,}")

    avg_records = canton_rooms + dur_recs
    print(f"\n  Total bfs_average_rents records to upsert: {len(avg_records):,}")

    # Normalise keys
    all_keys: set[str] = set()
    for r in avg_records:
        all_keys |= r.keys()
    for r in avg_records:
        for k in all_keys:
            r.setdefault(k, None)

    n_avg = batch_upsert(
        url=url, key=key, table=TABLE_AVG,
        records=avg_records,
        conflict_column=CONFLICT_AVG,
        schema=schema, batch_size=500,
    )
    print(f"  upserted to {TABLE_AVG}: {n_avg:,}")

    # ── Source 4: owner type → bfs_rents_by_owner_type ──
    print("\n[4] owner-type by canton (asset {})".format(ASSET_OWNER_TYPE_BY_CANTON))
    owner_bytes = _download_xlsx(ASSET_OWNER_TYPE_BY_CANTON)
    owner_recs = parse_owner_type(owner_bytes, ASSET_OWNER_TYPE_BY_CANTON)
    print(f"  parsed: {len(owner_recs):,}")

    if owner_recs:
        n_owner = batch_upsert(
            url=url, key=key, table=TABLE_OWNER,
            records=owner_recs,
            conflict_column=CONFLICT_OWNER,
            schema=schema, batch_size=500,
        )
        print(f"  upserted to {TABLE_OWNER}: {n_owner:,}")
    else:
        n_owner = 0
        print("  no owner-type rows parsed (skipping upsert)")

    print("\n" + "=" * 60)
    print(f"  IMPORT COMPLETE in {time.time()-t0:.1f}s")
    print(f"  bfs_average_rents:        {n_avg:,} upserted")
    print(f"  bfs_rents_by_owner_type:  {n_owner:,} upserted")
    print("=" * 60)


if __name__ == "__main__":
    main()

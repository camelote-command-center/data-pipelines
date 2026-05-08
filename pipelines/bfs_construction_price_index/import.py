#!/usr/bin/env python3
"""
BFS Construction Price Index (Baupreisindex / BAP) — Import Pipeline

Source:  BFS DAM, latest "Aktuelle Resultate pro Grossregion" xlsx asset
         (auto-discovered each run via the DAM list API filtered on inquiry=BAP).

         The xlsx contains 8 sheets — one per BFS Grande région (Schweiz +
         7 regions) — each with 16 object types (Baugewerbe Total, Hochbau,
         Tiefbau, Neubau Mehrfamilienhaus, Renovation Bürogebäude, ...). Each
         release is a snapshot for one publication month (e.g. October 2025),
         with weight %, index level, q-o-q variation %, y-o-y variation %.

         Snapshot semantics: current-only forward, no historical backfill
         (per Ilan's Q3 audit answer). Each semi-annual release adds a new
         period to the table.

Target:  bronze_ch.bfs_construction_price_index
Conflict key: (period, region_code, object_code)
Cadence: semi-annual (BFS publishes mid-April + mid-October)

Environment variables:
    RE_LLM_SUPABASE_URL              - re-LLM Supabase project URL (required)
    RE_LLM_SUPABASE_SERVICE_ROLE_KEY - service_role key (required)
    RE_LLM_SCHEMA                    - target schema (default: bronze_ch)
"""

from __future__ import annotations

import io
import os
import re
import sys

import openpyxl
import requests

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
from shared.supabase_client import batch_upsert  # noqa: E402

# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

DAM_LIST_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets"
DAM_MASTER_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/{asset_id}/master"

TABLE_NAME = "bfs_construction_price_index"

# Region code → human name (DE/EN canonical, full name kept in xlsx)
REGION_NAMES = {
    1: "Switzerland",
    2: "Lake Geneva region (VD/VS/GE)",
    3: "Espace Mittelland (BE/FR/SO/NE/JU)",
    4: "Northwestern Switzerland (BS/BL/AG)",
    5: "Zurich (ZH)",
    6: "Eastern Switzerland (GL/SH/AR/AI/SG/GR/TG)",
    7: "Central Switzerland (LU/UR/SZ/OW/NW/ZG)",
    8: "Ticino (TI)",
}

# German months → ISO month numbers, for parsing period labels in xlsx
DE_MONTHS = {
    "Januar": 1, "Februar": 2, "März": 3, "April": 4, "Mai": 5, "Juni": 6,
    "Juli": 7, "August": 8, "September": 9, "Oktober": 10, "November": 11,
    "Dezember": 12,
}

_PERIOD_DE_RE = re.compile(r"(Januar|Februar|März|April|Mai|Juni|Juli|August|September|Oktober|November|Dezember)\s+(\d{4})")


# ──────────────────────────────────────────────────────────────
# Source discovery
# ──────────────────────────────────────────────────────────────

def find_latest_grossregion_asset() -> tuple[int, str]:
    """Query DAM API for the most recent BAP 'Aktuelle Resultate pro Grossregion' xlsx.

    Returns (asset_id, title). The multilingual TABL variant is preferred
    (DE/FR/IT, single asset for all 3 languages). Newest = highest damId.

    DAM API caps page size at 20, so we paginate up to 5 pages (100 assets).
    The multilang asset for the latest semester typically appears within the
    first 100 results sorted by default ranking.
    """
    candidates: list[tuple[int, str]] = []
    for skip in range(0, 100, 20):
        r = requests.get(
            DAM_LIST_URL,
            params={"inquiry": "BAP", "size": 20, "skip": skip},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json().get("data", [])
        if not data:
            break
        for a in data:
            am = a.get("bfs", {}).get("articleModel", {}) or {}
            desc = a.get("description", {}) or {}
            titles = desc.get("titles", {}) or {}
            main = str(titles.get("main", ""))
            lang = str(desc.get("language", ""))
            if (am.get("code") == "TABL"
                    and "DE/FR/IT" in lang
                    and "Aktuelle Resultate pro Grossregion" in main):
                asset_id = a.get("ids", {}).get("damId")
                candidates.append((asset_id, main))

    if not candidates:
        raise RuntimeError(
            "No BAP 'Aktuelle Resultate pro Grossregion' multilang TABL xlsx "
            "found in DAM (searched first 100 BAP assets)."
        )

    candidates.sort(key=lambda x: -(x[0] or 0))
    return candidates[0]


# ──────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────

def _parse_period_label(label: str) -> str | None:
    """'Oktober 2025' → '2025-10'. Returns None if not parseable."""
    if not label:
        return None
    m = _PERIOD_DE_RE.search(str(label))
    if not m:
        return None
    month = DE_MONTHS.get(m.group(1))
    year = int(m.group(2))
    if month is None:
        return None
    return f"{year}-{month:02d}"


def parse_xlsx(xlsx_bytes: bytes) -> list[dict]:
    """Iterate sheets named '1'..'8', each a region. Yield long-format rows.

    Sheet layout (per region):
      r5  col 0 = '<REG_0X>'  col 1 = region label  col 5 = 'Basis Oktober 2020 = 100'
      r6  col 2 = 'Gewicht in %'  col 3 = 'Index'  col 4 = 'Veränderung in %...'
      r7  col 3 = period label (e.g. 'Oktober 2025')
          col 4 = q-o-q comparison period (e.g. 'April 2025')
          col 5 = y-o-y comparison period (e.g. 'Oktober 2024')
      r8..r23  data rows: col 0 = '<OBJ_NN>', col 1 = label, cols 2..5 = values
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)

    region_sheets = [str(i) for i in range(1, 9) if str(i) in wb.sheetnames]
    if not region_sheets:
        raise RuntimeError(f"No region sheets (1..8) in workbook; sheets={wb.sheetnames}")

    out: list[dict] = []

    for region_code_str in region_sheets:
        region_code = int(region_code_str)
        ws = wb[region_code_str]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 8:
            continue

        # Period header at row 7
        period = _parse_period_label(rows[7][3] if len(rows[7]) > 3 else None)
        comp_qoq = str(rows[7][4]) if len(rows[7]) > 4 and rows[7][4] else None
        comp_yoy = str(rows[7][5]) if len(rows[7]) > 5 and rows[7][5] else None

        # Base period from row 5 col 5 (e.g. "Basis Oktober 2020 = 100")
        base_label = str(rows[5][5]) if len(rows[5]) > 5 and rows[5][5] else ""
        m = re.search(r"Basis\s+(.+?)\s*=", base_label)
        base_period = m.group(1).strip() if m else None

        if period is None:
            print(f"  Sheet {region_code_str}: could not parse period — skipping")
            continue

        # Data rows: 8..23 (16 object types)
        for r_idx in range(8, min(24, len(rows))):
            row = rows[r_idx]
            if len(row) < 6:
                continue
            obj_code_raw = row[0]
            obj_label = row[1]
            if not obj_code_raw or not obj_label:
                continue
            obj_code = str(obj_code_raw).strip().strip("<>")  # 'OBJ_02'

            try:
                weight = float(row[2]) if row[2] is not None and row[2] != "" else None
                index_value = float(row[3]) if row[3] is not None and row[3] != "" else None
                var_qoq = float(row[4]) if row[4] is not None and row[4] != "" else None
                var_yoy = float(row[5]) if row[5] is not None and row[5] != "" else None
            except (TypeError, ValueError):
                continue

            if index_value is None:
                continue

            out.append({
                "period": period,
                "region_code": region_code,
                "region_name": REGION_NAMES.get(region_code),
                "object_code": obj_code,
                "object_label": str(obj_label).strip(),
                "weight_pct": weight,
                "index_value": index_value,
                "base_period": base_period,
                "variation_qoq_pct": var_qoq,
                "variation_qoq_compared_to": comp_qoq,
                "variation_yoy_pct": var_yoy,
                "variation_yoy_compared_to": comp_yoy,
            })

    return out


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def _required(name: str) -> str:
    v = os.environ.get(name, "")
    if not v:
        print(f"ERROR: env var {name} required")
        sys.exit(1)
    return v


def main():
    url = _required("RE_LLM_SUPABASE_URL")
    key = _required("RE_LLM_SUPABASE_SERVICE_ROLE_KEY")
    schema = os.environ.get("RE_LLM_SCHEMA", "bronze_ch")

    print("=" * 60)
    print("  BFS Construction Price Index Import Pipeline")
    print(f"  Target: {schema}.{TABLE_NAME} on {url}")
    print("=" * 60)

    print("\n[1/3] Discovering latest BAP 'Aktuelle Resultate' asset…")
    asset_id, title = find_latest_grossregion_asset()
    print(f"  asset_id = {asset_id}")
    print(f"  title    = {title}")

    print("\n[2/3] Downloading xlsx…")
    r = requests.get(
        DAM_MASTER_URL.format(asset_id=asset_id),
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    r.raise_for_status()
    print(f"  bytes received: {len(r.content):,}")

    rows = parse_xlsx(r.content)
    print(f"  parsed rows: {len(rows):,}")
    if rows:
        print(f"  sample period(s): {sorted(set(r['period'] for r in rows))}")

    if not rows:
        print("ERROR: no rows parsed — refusing to upsert empty payload.")
        sys.exit(1)

    print("\n[3/3] Upserting …")
    n = batch_upsert(
        url=url,
        key=key,
        table=TABLE_NAME,
        records=rows,
        conflict_column="period,region_code,object_code",
        schema=schema,
        batch_size=500,
    )
    print(f"  upserted: {n}")

    print("\n" + "=" * 60)
    print("  IMPORT COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()

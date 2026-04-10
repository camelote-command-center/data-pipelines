#!/usr/bin/env python3
"""
BFS Population by Commune — Import Pipeline

Downloads commune-level population data from the Swiss Federal Statistical
Office (BFS) via BFS DAM API and upserts into bronze.bfs_population
on lamap_db.

Source:  BFS DAM API / opendata.swiss (fallback)
         Dataset: Ständige Wohnbevölkerung nach Gemeinde
         XLSX/CSV download

Target:  bronze.bfs_population
Conflict: year,bfs_commune_number

DATA SAFETY:
    - UPSERT only. Never truncates or deletes.
    - Row count should only go UP or stay the same.

Environment variables:
    LAMAP_SUPABASE_URL          - Lamap Supabase project URL (required)
    LAMAP_SUPABASE_SERVICE_KEY  - service_role key (required)
    LAMAP_SCHEMA                - target schema (default: bronze)
    CAMELOTE_SUPABASE_URL       - Command center URL (optional, for metadata)
    CAMELOTE_SUPABASE_KEY       - Command center key (optional)
"""

import csv
import io
import os
import re
import sys
import time

import openpyxl
import requests

# Add repo root to path so we can import shared/
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
from shared.supabase_client import batch_upsert
from shared.freshness import get_dataset_meta, update_dataset_meta


# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

# Direct BFS ASSET URLs (PRIMARY) — BFS population by commune (XLSX)
# Asset IDs change over time; we try several from newest to oldest.
BFS_ASSET_IDS = [
    "36451447",  # 2025 provisional population by commune
    "34447410",  # 2024 provisional population by commune
    "32007762",  # older asset (may be gone)
]
BFS_ASSET_URL_TEMPLATE = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/{}/master"

# opendata.swiss dataset search (FALLBACK) — searches for BFS population data
OPENDATA_SEARCH_URL = (
    "https://opendata.swiss/api/3/action/package_search"
    "?q=bevoelkerung+gemeinde+bfs&rows=10"
)

# Reject URLs from cantonal portals (they return partial/unrelated data)
REJECTED_DOMAINS = {
    "data.zg.ch", "data.bs.ch", "data.bl.ch", "data.be.ch", "data.zh.ch",
    "daten.statistik.zh.ch",
}

TABLE = "bfs_population"
CONFLICT_COLUMN = "year,bfs_commune_number"
BATCH_SIZE = 500
DATASET_CODE = "ch_bfs_population"

# Canton abbreviation → code (for "- Zürich" → canton tracking)
CANTON_NAME_TO_CODE = {
    "zürich": "ZH", "bern": "BE", "bern / berne": "BE", "luzern": "LU", "uri": "UR",
    "schwyz": "SZ", "obwalden": "OW", "nidwalden": "NW", "glarus": "GL",
    "zug": "ZG", "freiburg": "FR", "fribourg": "FR", "freiburg / fribourg": "FR",
    "fribourg / freiburg": "FR",
    "solothurn": "SO", "basel-stadt": "BS", "basel-landschaft": "BL",
    "schaffhausen": "SH", "appenzell ausserrhoden": "AR", "appenzell a.rh.": "AR",
    "appenzell innerrhoden": "AI", "appenzell i.rh.": "AI",
    "st. gallen": "SG", "st.gallen": "SG", "graubünden": "GR",
    "graubünden / grigioni / grischun": "GR",
    "aargau": "AG", "thurgau": "TG", "tessin": "TI", "ticino": "TI",
    "waadt": "VD", "vaud": "VD", "wallis": "VS", "valais": "VS",
    "wallis / valais": "VS", "valais / wallis": "VS",
    "neuenburg": "NE", "neuchâtel": "NE", "genf": "GE", "genève": "GE",
    "jura": "JU",
}

# Column mapping for CSV fallback
COLUMN_MAP = {
    "jahr": "year", "annee": "year", "year": "year", "stichtag": "year",
    "gemeinde_nummer": "bfs_commune_number", "gemeindenummer": "bfs_commune_number",
    "gemeinde-nr.": "bfs_commune_number", "bfs_nr": "bfs_commune_number",
    "bfs_commune_number": "bfs_commune_number", "no_commune": "bfs_commune_number",
    "commune_id": "bfs_commune_number",
    "gemeindename": "commune_name", "gemeinde": "commune_name",
    "commune": "commune_name", "commune_name": "commune_name", "nom_commune": "commune_name",
    "kanton": "canton_code", "kantonskuerzel": "canton_code",
    "canton": "canton_code", "canton_code": "canton_code", "kt": "canton_code",
    "bevoelkerung": "total_population", "total": "total_population",
    "total_population": "total_population", "einwohner": "total_population",
    "population_totale": "total_population", "wohnbevoelkerung": "total_population",
    "schweizer": "swiss_nationals", "swiss_nationals": "swiss_nationals",
    "schweizer_innen": "swiss_nationals", "suisses": "swiss_nationals",
    "auslaender": "foreign_nationals", "auslaender_innen": "foreign_nationals",
    "foreign_nationals": "foreign_nationals", "etrangers": "foreign_nationals",
}


# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────

def _resource_name_str(name) -> str:
    """Extract a plain string from a resource name that may be a multilingual dict."""
    if isinstance(name, dict):
        return name.get("de") or name.get("en") or name.get("fr") or next(iter(name.values()), "")
    if isinstance(name, str):
        return name
    return ""


def _is_rejected_url(url: str) -> bool:
    """Reject URLs from cantonal portals that return partial/unrelated data."""
    from urllib.parse import urlparse
    domain = urlparse(url).hostname or ""
    return domain in REJECTED_DOMAINS


def _safe_int(val) -> int | None:
    """Convert a cell value to int, handling floats and strings."""
    if val is None:
        return None
    try:
        if isinstance(val, float):
            return int(val)
        return int(str(val).strip().replace("'", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def get_row_count(url: str, key: str, schema: str) -> int | None:
    """Get current row count via PostgREST HEAD request."""
    endpoint = f"{url.rstrip('/')}/rest/v1/{TABLE}?select=count"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Prefer": "count=exact",
    }
    if schema and schema != "public":
        headers["Accept-Profile"] = schema
    try:
        r = requests.head(endpoint, headers=headers, timeout=30)
        cr = r.headers.get("content-range", "")
        if "/" in cr:
            return int(cr.split("/")[1])
    except Exception as e:
        print(f"  Warning: could not get row count: {e}")
    return None


# ──────────────────────────────────────────────────────────────
# XLSX Parser (BFS DAM API format)
# ──────────────────────────────────────────────────────────────

def download_binary(url: str, retries: int = 3) -> bytes:
    """Download file content as bytes with retry on timeout."""
    for attempt in range(1, retries + 1):
        try:
            print(f"  Downloading: {url}" + (f" (attempt {attempt})" if attempt > 1 else ""))
            r = requests.get(url, timeout=120)
            r.raise_for_status()
            print(f"  Downloaded {len(r.content):,} bytes")
            return r.content
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            if attempt < retries:
                wait = attempt * 10
                print(f"  Timeout/connection error, retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise


def parse_xlsx_to_records(xlsx_bytes: bytes) -> list[dict]:
    """
    Parse BFS population XLSX.

    Single sheet with structure:
      Row 0: title containing date "am 31.12.YYYY"
      Row 1: subtitle
      Row 2: column group headers (Total, Schweizer, Ausländer)
      Row 3: sub-headers (Total, Mann, Frau for each group)
      Row 4: Switzerland total
      Row 5+: Data rows:
        "- Zürich" → canton header
        ">> Bezirk Name" → district header
        "......0001 Commune Name" → commune data

    Columns: [Region, Total-Total, Total-M, Total-F, CH-Total, CH-M, CH-F, Ausl-Total, Ausl-M, Ausl-F]
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records = []

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 5:
        print("  ERROR: XLSX has fewer than 5 rows")
        wb.close()
        return []

    # Extract year from title row (e.g., "am 31.12.2025")
    title = str(rows[0][1] or "") + " " + str(rows[0][0] or "")
    year_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', title)
    if year_match:
        year = int(year_match.group(3))
    else:
        # Try to find any 4-digit year
        year_match = re.search(r'(20\d{2})', title)
        if year_match:
            year = int(year_match.group(1))
        else:
            print(f"  ERROR: Could not extract year from title: {title}")
            wb.close()
            return []

    print(f"  Detected year: {year}")

    # Column indices (0-based):
    # 0: Region name
    # 1: Total population (Total-Total)
    # 4: Swiss nationals (CH-Total)
    # 7: Foreign nationals (Ausl-Total)
    COL_TOTAL = 1
    COL_SWISS = 4
    COL_FOREIGN = 7

    current_canton = None

    for row in rows:
        if not row or not row[0]:
            continue

        cell = str(row[0]).strip()

        # Canton header: "- Zürich"
        if cell.startswith("- "):
            canton_name = cell[2:].strip().lower()
            # Strip footnote markers
            canton_name = re.sub(r'\s*\d+\)\s*$', '', canton_name).strip()
            current_canton = CANTON_NAME_TO_CODE.get(canton_name)
            if not current_canton:
                print(f"  Warning: unknown canton '{cell[2:].strip()}'")
            continue

        # Commune data: "......0001 Commune Name"
        m = re.match(r'\.{4,}(\d+)\s+(.*)', cell)
        if not m:
            continue

        bfs_number = int(m.group(1))
        commune_name = m.group(2).strip()
        # Strip footnote markers from commune name
        commune_name = re.sub(r'\s*\d+\)\s*$', '', commune_name).strip()

        record = {
            "year": year,
            "bfs_commune_number": bfs_number,
            "commune_name": commune_name,
            "canton_code": current_canton,
            "total_population": _safe_int(row[COL_TOTAL]) if COL_TOTAL < len(row) else None,
            "swiss_nationals": _safe_int(row[COL_SWISS]) if COL_SWISS < len(row) else None,
            "foreign_nationals": _safe_int(row[COL_FOREIGN]) if COL_FOREIGN < len(row) else None,
        }
        records.append(record)

    wb.close()
    return records


# ──────────────────────────────────────────────────────────────
# CSV Parser (opendata.swiss fallback)
# ──────────────────────────────────────────────────────────────

def find_csv_url_opendata() -> str | None:
    """Search opendata.swiss for the BFS population CSV download URL."""
    try:
        r = requests.get(OPENDATA_SEARCH_URL, timeout=30)
        r.raise_for_status()
        data = r.json()
        results = data.get("result", {}).get("results", [])

        for dataset in results:
            org = (dataset.get("organization", {}) or {}).get("name", "")
            for resource in dataset.get("resources", []):
                fmt = (resource.get("format") or "").lower()
                url = resource.get("url", "")
                name = _resource_name_str(resource.get("name")).lower()
                if fmt in ("csv", "text/csv") and (
                    "gemeinde" in name or "commune" in name or "population" in name
                    or "bevoelkerung" in name
                ):
                    if _is_rejected_url(url):
                        print(f"  Skipped (cantonal portal): {url}")
                        continue
                    print(f"  Found CSV: {_resource_name_str(resource.get('name'))} (org: {org})")
                    return url

        # Broader search: any CSV resource (still filtering cantonal portals)
        for dataset in results:
            for resource in dataset.get("resources", []):
                fmt = (resource.get("format") or "").lower()
                url = resource.get("url", "")
                if fmt in ("csv", "text/csv") and not _is_rejected_url(url):
                    print(f"  Found CSV (broad match): {_resource_name_str(resource.get('name'))}")
                    return url

    except Exception as e:
        print(f"  Warning: opendata.swiss search failed: {e}")
    return None


def download_csv(csv_url: str) -> str:
    """Download CSV content as text."""
    print(f"  Downloading: {csv_url}")
    r = requests.get(csv_url, timeout=120)
    r.raise_for_status()
    text = r.content.decode("utf-8-sig")
    print(f"  Downloaded {len(text):,} characters")
    return text


def detect_delimiter(text: str) -> str:
    """Detect CSV delimiter (semicolon or comma)."""
    first_line = text.split("\n")[0]
    if first_line.count(";") > first_line.count(","):
        return ";"
    return ","


def parse_csv_to_records(text: str) -> list[dict]:
    """Parse CSV text into table records using column mapping."""
    delimiter = detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    header_map = {}
    if reader.fieldnames:
        for h in reader.fieldnames:
            key = h.strip().lower().replace(" ", "_").replace("-", "_")
            if key in COLUMN_MAP:
                header_map[h] = COLUMN_MAP[key]

    if not header_map:
        print(f"  WARNING: Could not map any CSV headers")
        print(f"  Available headers: {reader.fieldnames}")
        return []

    print(f"  Mapped columns: {header_map}")

    records = []
    for row in reader:
        record = {}
        for csv_col, table_col in header_map.items():
            val = row.get(csv_col, "").strip()
            record[table_col] = val if val else None

        if not record.get("bfs_commune_number") or not record.get("year"):
            continue

        try:
            record["bfs_commune_number"] = int(record["bfs_commune_number"])
        except (ValueError, TypeError):
            continue

        year_val = record.get("year", "")
        if year_val:
            try:
                if "-" in str(year_val):
                    record["year"] = int(str(year_val).split("-")[0])
                else:
                    record["year"] = int(year_val)
            except (ValueError, TypeError):
                continue

        for col in ("total_population", "swiss_nationals", "foreign_nationals"):
            if record.get(col) is not None:
                try:
                    record[col] = int(record[col])
                except (ValueError, TypeError):
                    record[col] = None

        records.append(record)

    return records


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def main():
    lamap_url = os.environ.get("LAMAP_SUPABASE_URL", "")
    lamap_key = os.environ.get("LAMAP_SUPABASE_SERVICE_KEY", "")
    lamap_schema = os.environ.get("LAMAP_SCHEMA", "bronze")
    camelote_url = os.environ.get("CAMELOTE_SUPABASE_URL", "")
    camelote_key = os.environ.get("CAMELOTE_SUPABASE_KEY", "")

    if not lamap_url or not lamap_key:
        print("ERROR: LAMAP_SUPABASE_URL and LAMAP_SUPABASE_SERVICE_KEY are required")
        sys.exit(1)

    print("=" * 60)
    print("  BFS Population by Commune Pipeline")
    print(f"  Target: {lamap_schema}.{TABLE}")
    print("=" * 60)

    # ── Row count BEFORE ──
    rows_before = get_row_count(lamap_url, lamap_key, lamap_schema)
    print(f"\n  Rows before: {rows_before:,}" if rows_before is not None else "\n  Rows before: unknown")

    # ── Download & Parse ──
    start = time.time()
    records = []
    source_url = None

    # Strategy 1: BFS DAM API (primary — XLSX with commune data)
    print("\n  Trying BFS DAM API assets (XLSX)...")
    for asset_id in BFS_ASSET_IDS:
        asset_url = BFS_ASSET_URL_TEMPLATE.format(asset_id)
        print(f"  Trying asset {asset_id}...")
        try:
            data = download_binary(asset_url)
            content_type = ""
            try:
                r = requests.head(asset_url, timeout=10)
                content_type = r.headers.get("content-type", "")
            except Exception:
                pass

            # Detect format: XLSX or CSV
            if (content_type and "spreadsheet" in content_type) or data[:4] == b'PK\x03\x04':
                print("  Detected XLSX format, parsing...")
                records = parse_xlsx_to_records(data)
            else:
                print("  Detected CSV format, parsing...")
                text = data.decode("utf-8-sig")
                records = parse_csv_to_records(text)

            if records:
                source_url = asset_url
                print(f"  Success: asset {asset_id} → {len(records):,} records")
                break
            else:
                print(f"  Asset {asset_id}: downloaded but parsed 0 records, trying next...")
        except requests.exceptions.HTTPError:
            print(f"  Asset {asset_id} not available, trying next...")
            continue
        except Exception as e:
            print(f"  Asset {asset_id} error: {e}, trying next...")
            continue

    # Strategy 2: opendata.swiss search (fallback — CSV)
    if not records:
        print("\n  BFS DAM assets unavailable, searching opendata.swiss...")
        csv_url = find_csv_url_opendata()
        if csv_url:
            try:
                text = download_csv(csv_url)
                print("\n  Parsing CSV...")
                records = parse_csv_to_records(text)
                if records:
                    source_url = csv_url
            except requests.exceptions.HTTPError as e:
                print(f"  Warning: opendata.swiss URL failed ({e})")

    print(f"  Parsed {len(records):,} records total")

    if not records:
        print("  ERROR: Could not download or parse population data from any source")
        sys.exit(1)

    # Normalise keys
    all_keys = set()
    for rec in records:
        all_keys |= rec.keys()
    for rec in records:
        for k in all_keys:
            rec.setdefault(k, None)

    # ── Upsert ──
    print(f"\n  Upserting {len(records):,} records...")
    upserted = batch_upsert(
        url=lamap_url,
        key=lamap_key,
        table=TABLE,
        records=records,
        conflict_column=CONFLICT_COLUMN,
        schema=lamap_schema,
        batch_size=BATCH_SIZE,
    )

    elapsed = time.time() - start

    # ── Row count AFTER ──
    rows_after = get_row_count(lamap_url, lamap_key, lamap_schema)

    # ── Summary ──
    print(f"\n{'=' * 60}")
    print("  IMPORT COMPLETE")
    print(f"  Source:          {source_url}")
    print(f"  Records parsed:  {len(records):,}")
    print(f"  Rows upserted:   {upserted:,}")
    print(f"  Rows before:     {rows_before:,}" if rows_before is not None else "  Rows before:     unknown")
    print(f"  Rows after:      {rows_after:,}" if rows_after is not None else "  Rows after:      unknown")
    if rows_before is not None and rows_after is not None:
        delta = rows_after - rows_before
        print(f"  Net new:         {delta:,}")
    print(f"  Duration:        {elapsed:.1f}s")
    print("=" * 60)

    if upserted == 0:
        print("  FAILED: Zero rows upserted!")
        sys.exit(1)

    # ── Update dataset metadata ──
    update_dataset_meta(
        camelote_url, camelote_key, DATASET_CODE,
        record_count=rows_after,
        status="active",
    )


if __name__ == "__main__":
    main()

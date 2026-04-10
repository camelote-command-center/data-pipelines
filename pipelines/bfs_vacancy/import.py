#!/usr/bin/env python3
"""
BFS Vacancy Rates — Import Pipeline

Downloads vacancy rate data (Leerwohnungszaehlung) from the Swiss Federal
Statistical Office (BFS) via BFS DAM API and upserts into
bronze.bfs_vacancy_rates on lamap_db.

Source:  BFS DAM API / opendata.swiss (fallback)
         Dataset: Leer stehende Wohnungen nach Kantonen (T 09.03.04.03)
         XLSX download from BFS DAM API

Target:  bronze.bfs_vacancy_rates
Conflict: year,canton_code

Currently ~700+ rows at canton level (26 cantons × ~26 years).
BFS publishes annually (usually September).

DATA SAFETY:
    - UPSERT only. Never truncates or deletes.
    - Row count should only go UP or stay the same.

Environment variables:
    LAMAP_SUPABASE_URL          - Lamap Supabase project URL (required)
    LAMAP_SUPABASE_SERVICE_KEY  - service_role key (required)
    LAMAP_SCHEMA                - target schema (default: bronze)
    CAMELOTE_SUPABASE_URL       - Command center URL (optional, for metadata)
    CAMELOTE_SUPABASE_KEY       - Command center key (optional)
"""

import csv
import io
import os
import re
import sys
import time

import openpyxl
import requests

# Add repo root to path so we can import shared/
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
from shared.supabase_client import batch_upsert
from shared.freshness import get_dataset_meta, update_dataset_meta


# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

# Direct BFS ASSET URLs (PRIMARY) — national vacancy data by canton (XLSX)
# Dataset: "Leer stehende Wohnungen nach Kantonen" (T 09.03.04.03)
BFS_ASSET_IDS = [
    "36153026",  # DE — 2025 edition, multi-sheet XLSX
    "36153028",  # EN — same data in English
]
BFS_ASSET_URL_TEMPLATE = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/{}/master"

# opendata.swiss search (FALLBACK)
OPENDATA_SEARCH_URL = (
    "https://opendata.swiss/api/3/action/package_search"
    "?q=leerwohnungsziffer+leerwohnung+bfs&rows=10"
)

# Reject URLs from cantonal portals (they return partial/unrelated data)
REJECTED_DOMAINS = {"data.zg.ch", "data.bs.ch", "data.bl.ch", "data.be.ch", "data.zh.ch"}

TABLE = "bfs_vacancy_rates"
CONFLICT_COLUMN = "year,canton_code"
BATCH_SIZE = 500
DATASET_CODE = "ch_bfs_vacancy_rates"

# Swiss canton name → code mapping (German names as used in BFS XLSX)
CANTON_NAME_TO_CODE = {
    "zürich": "ZH", "bern": "BE", "luzern": "LU", "uri": "UR",
    "schwyz": "SZ", "obwalden": "OW", "nidwalden": "NW", "glarus": "GL",
    "zug": "ZG", "freiburg": "FR", "fribourg": "FR",
    "solothurn": "SO", "basel-stadt": "BS", "basel-landschaft": "BL",
    "schaffhausen": "SH", "appenzell a.rh.": "AR", "appenzell ausserrhoden": "AR",
    "appenzell i.rh.": "AI", "appenzell innerrhoden": "AI",
    "st. gallen": "SG", "st.gallen": "SG", "graubünden": "GR",
    "aargau": "AG", "thurgau": "TG", "tessin": "TI", "ticino": "TI",
    "waadt": "VD", "vaud": "VD", "wallis": "VS", "valais": "VS",
    "neuenburg": "NE", "neuchâtel": "NE", "genf": "GE", "genève": "GE",
    "jura": "JU",
    # English names (for EN asset)
    "zurich": "ZH", "berne": "BE", "lucerne": "LU",
    "freiburg/fribourg": "FR", "basle-city": "BS", "basle-country": "BL",
    "appenzell outer rhodes": "AR", "appenzell inner rhodes": "AI",
    "st.gall": "SG", "grisons": "GR", "ticino": "TI",
    "vaud": "VD", "valais": "VS", "neuchatel": "NE", "geneva": "GE",
}

SWISS_CANTONS = set(CANTON_NAME_TO_CODE.values())

# Column mapping for CSV fallback
COLUMN_MAP = {
    # Year
    "jahr": "year",
    "annee": "year",
    "year": "year",
    "stichtag": "year",
    # Canton
    "kanton": "canton_code",
    "kantonskuerzel": "canton_code",
    "canton": "canton_code",
    "canton_code": "canton_code",
    "kt": "canton_code",
    "kanton_kuerzel": "canton_code",
    # Commune (for commune-level data if available)
    "gemeinde_nummer": "bfs_commune_number",
    "gemeindenummer": "bfs_commune_number",
    "bfs_nr": "bfs_commune_number",
    "bfs_commune_number": "bfs_commune_number",
    "no_commune": "bfs_commune_number",
    # Commune name
    "gemeindename": "commune_name",
    "gemeinde": "commune_name",
    "commune": "commune_name",
    "commune_name": "commune_name",
    # Dwelling counts
    "wohnungen_total": "total_dwellings",
    "total_dwellings": "total_dwellings",
    "gesamtzahl": "total_dwellings",
    "wohnungsbestand": "total_dwellings",
    "total_logements": "total_dwellings",
    # Vacant
    "leerwohnungen": "vacant_dwellings",
    "leer": "vacant_dwellings",
    "vacant_dwellings": "vacant_dwellings",
    "leerstehend": "vacant_dwellings",
    "logements_vacants": "vacant_dwellings",
    # Vacancy rate
    "leerwohnungsziffer": "vacancy_rate_pct",
    "vacancy_rate_pct": "vacancy_rate_pct",
    "quote": "vacancy_rate_pct",
    "taux_vacance": "vacancy_rate_pct",
    "leerwohnungsquote": "vacancy_rate_pct",
}

# Expected columns in a valid BFS vacancy CSV (for CSV fallback validation)
EXPECTED_VACANCY_HEADERS = {
    "leerwohnungsziffer", "leerwohnungsquote", "vacancy_rate_pct",
    "taux_vacance", "quote", "leerwohnungen", "logements_vacants",
    "vacant_dwellings", "leerstehend",
}


# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────

def _resource_name_str(name) -> str:
    """Extract a plain string from a resource name that may be a multilingual dict."""
    if isinstance(name, dict):
        return name.get("de") or name.get("en") or name.get("fr") or next(iter(name.values()), "")
    if isinstance(name, str):
        return name
    return ""


def _is_rejected_url(url: str) -> bool:
    """Reject URLs from cantonal portals that return partial/unrelated data."""
    from urllib.parse import urlparse
    domain = urlparse(url).hostname or ""
    return domain in REJECTED_DOMAINS


def get_row_count(url: str, key: str, schema: str) -> int | None:
    endpoint = f"{url.rstrip('/')}/rest/v1/{TABLE}?select=count"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Prefer": "count=exact",
    }
    if schema and schema != "public":
        headers["Accept-Profile"] = schema
    try:
        r = requests.head(endpoint, headers=headers, timeout=30)
        cr = r.headers.get("content-range", "")
        if "/" in cr:
            return int(cr.split("/")[1])
    except Exception as e:
        print(f"  Warning: could not get row count: {e}")
    return None


def _extract_canton_code(region_name: str) -> str | None:
    """Extract 2-letter canton code from a region name like 'Zürich', 'ZH', 'Zürich (ZH)', or 'Bern 5)'."""
    if not region_name:
        return None
    stripped = region_name.strip()
    # Try parenthesised code first: "Zürich (ZH)"
    m = re.search(r'\(([A-Z]{2})\)', stripped)
    if m and m.group(1) in SWISS_CANTONS:
        return m.group(1)
    # Strip footnote markers like "5)" or " 5)"
    cleaned = re.sub(r'\s*\d+\)\s*$', '', stripped).strip()
    # Try bare 2-letter code
    if len(cleaned) == 2 and cleaned.upper() in SWISS_CANTONS:
        return cleaned.upper()
    # Try name lookup
    return CANTON_NAME_TO_CODE.get(cleaned.lower())


def _safe_int(val) -> int | None:
    """Convert a cell value to int, handling floats and strings."""
    if val is None:
        return None
    try:
        if isinstance(val, float):
            return int(val)
        return int(str(val).strip().replace("'", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _safe_float(val) -> float | None:
    """Convert a cell value to float."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ──────────────────────────────────────────────────────────────
# XLSX Parser (BFS DAM API format)
# ──────────────────────────────────────────────────────────────

def download_xlsx(url: str) -> bytes:
    """Download XLSX content as bytes."""
    print(f"  Downloading: {url}")
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    print(f"  Downloaded {len(r.content):,} bytes")
    return r.content


def parse_xlsx_to_records(xlsx_bytes: bytes) -> list[dict]:
    """
    Parse BFS vacancy XLSX (multi-sheet, one sheet per year).

    Each sheet is named with a year (e.g. "2025(base StatBL2024)").
    Structure per sheet:
      Row 0: title
      Rows 1-4: multi-row headers (Wohnungsbestand, Leerwohnungen by rooms, Total, ..., Leerwohnungsziffer)
      Row 5: "Total" (Switzerland total)
      Rows 6+: Grossregionen and Cantons intermixed
      Footer rows: footnotes

    We extract canton-level rows (identified by canton name lookup).
    Wohnungsbestand is column 1, vacant Total is column 8, Leerwohnungsziffer is the last numeric column.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        # Try to extract year from sheet name
        year_match = re.search(r'(\d{4})', sheet_name)
        if not year_match:
            print(f"  Skipping sheet: {sheet_name} (no year)")
            continue

        year = int(year_match.group(1))
        if year < 1980 or year > 2100:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 6:
            continue

        # Find column indices by scanning header rows and the "Total" data row
        col_total_dwellings = None  # Wohnungsbestand (may not exist in older sheets)
        col_vacant_total = None     # Total vacant dwellings
        col_vacancy_rate = None     # Leerwohnungsziffer

        # Scan header rows for known labels
        for i, row in enumerate(rows[:10]):
            row_strs = [str(c or "").lower().strip() for c in row]
            combined = " ".join(row_strs)
            for j, cell in enumerate(row_strs):
                if "wohnungsbestand" in cell or "housing stock" in cell:
                    col_total_dwellings = j
                if "leerwohnungsziffer" in cell or "leerwohnungs-" in cell or "vacancy rate" in cell:
                    col_vacancy_rate = j
                # "ziffer" on a separate row (multi-row header)
                if cell.startswith("ziffer") and col_vacancy_rate is None:
                    col_vacancy_rate = j

        # Find the "Total" header for vacant dwellings
        for i in range(min(10, len(rows))):
            row_strs = [str(c or "").lower().strip() for c in rows[i]]
            for j, cell in enumerate(row_strs):
                if cell == "total" and j > 1:  # Skip first column
                    col_vacant_total = j
                    break
            if col_vacant_total:
                break

        # If we still don't have vacancy_rate, detect from the "Total" (Switzerland) data row
        # by finding the last small numeric value (vacancy rates are typically 0.3-5.0%)
        total_row = None
        for i, row in enumerate(rows[5:20], start=5):
            region = str(row[0] or "").strip().lower() if row[0] else ""
            if region == "total":
                total_row = row
                break

        if total_row and col_vacancy_rate is None:
            for j in range(len(total_row) - 1, 0, -1):
                val = _safe_float(total_row[j])
                if val is not None and 0 < val < 10:
                    col_vacancy_rate = j
                    break

        # Also detect col_vacant_total from the Total row if header scan missed it
        if total_row and col_vacant_total is None:
            # Find a column with a large value that looks like total vacant dwellings
            # It's the largest value before the subcategory columns
            best_j = None
            best_val = 0
            start_col = (col_total_dwellings or 0) + 1
            end_col = (col_vacancy_rate or len(total_row)) - 1
            for j in range(start_col, end_col):
                val = _safe_int(total_row[j])
                if val is not None and val > best_val:
                    best_val = val
                    best_j = j
            if best_j:
                col_vacant_total = best_j

        # Parse all data rows
        canton_count = 0
        for row in rows:
            if not row or not row[0]:
                continue

            region = str(row[0]).strip()
            if not region or region.startswith("1)") or region.startswith("2)") or "©" in region:
                continue

            canton_code = _extract_canton_code(region)
            if not canton_code:
                continue

            record = {
                "year": year,
                "canton_code": canton_code,
                "total_dwellings": _safe_int(row[col_total_dwellings]) if col_total_dwellings is not None and col_total_dwellings < len(row) else None,
                "vacant_dwellings": _safe_int(row[col_vacant_total]) if col_vacant_total is not None and col_vacant_total < len(row) else None,
                "vacancy_rate_pct": _safe_float(row[col_vacancy_rate]) if col_vacancy_rate is not None and col_vacancy_rate < len(row) else None,
                "source": "BFS Leerwohnungszaehlung",
            }
            records.append(record)
            canton_count += 1

        if canton_count > 0:
            print(f"  Sheet {sheet_name}: {canton_count} cantons")
        else:
            print(f"  Sheet {sheet_name}: 0 cantons (check format)")

    wb.close()
    return records


# ──────────────────────────────────────────────────────────────
# CSV Parser (opendata.swiss fallback)
# ──────────────────────────────────────────────────────────────

def _validate_vacancy_csv(csv_url: str) -> bool:
    """Download the first few KB of a CSV and check it has vacancy-related columns."""
    try:
        r = requests.get(csv_url, timeout=30, stream=True)
        r.raise_for_status()
        chunk = next(r.iter_content(4096, decode_unicode=True), "")
        r.close()
        if not chunk:
            return False
        first_line = chunk.split("\n")[0].lower()
        normalised = first_line.replace(" ", "_").replace("-", "_").replace('"', '')
        # Must have vacancy columns AND canton/region column
        has_vacancy = any(h in normalised for h in EXPECTED_VACANCY_HEADERS)
        has_canton = any(k in normalised for k in ("kanton", "canton", "kt", "kantonskuerzel"))
        return has_vacancy and has_canton
    except Exception as e:
        print(f"  Warning: could not validate CSV at {csv_url}: {e}")
        return False


def find_csv_url_opendata() -> str | None:
    """Search opendata.swiss for the BFS vacancy CSV download URL."""
    try:
        r = requests.get(OPENDATA_SEARCH_URL, timeout=30)
        r.raise_for_status()
        data = r.json()
        results = data.get("result", {}).get("results", [])

        csv_candidates = []
        for dataset in results:
            dataset_title = _resource_name_str(dataset.get("title")).lower() if isinstance(dataset.get("title"), dict) else (dataset.get("title") or "").lower()
            for resource in dataset.get("resources", []):
                fmt = (resource.get("format") or "").lower()
                if fmt in ("csv", "text/csv"):
                    url = resource.get("url", "")
                    if _is_rejected_url(url):
                        print(f"  Skipped (cantonal portal): {url}")
                        continue
                    name = _resource_name_str(resource.get("name")).lower()
                    csv_candidates.append((name, dataset_title, url))

        for name, dataset_title, url in csv_candidates:
            print(f"  Checking CSV: {name} (dataset: {dataset_title})")
            if _validate_vacancy_csv(url):
                print(f"  Validated CSV with vacancy + canton columns: {name}")
                return url
            else:
                print(f"  Skipped (missing vacancy or canton columns): {name}")

    except Exception as e:
        print(f"  Warning: opendata.swiss search failed: {e}")
    return None


def download_csv(csv_url: str) -> str:
    print(f"  Downloading: {csv_url}")
    r = requests.get(csv_url, timeout=120)
    r.raise_for_status()
    text = r.content.decode("utf-8-sig")
    print(f"  Downloaded {len(text):,} characters")
    return text


def detect_delimiter(text: str) -> str:
    first_line = text.split("\n")[0]
    if first_line.count(";") > first_line.count(","):
        return ";"
    return ","


def parse_csv_to_records(text: str) -> list[dict]:
    delimiter = detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    header_map = {}
    if reader.fieldnames:
        for h in reader.fieldnames:
            key = h.strip().lower().replace(" ", "_").replace("-", "_")
            if key in COLUMN_MAP:
                header_map[h] = COLUMN_MAP[key]

    if not header_map:
        print(f"  WARNING: Could not map any CSV headers")
        print(f"  Available headers: {reader.fieldnames}")
        return []

    print(f"  Mapped columns: {header_map}")

    records = []
    for row in reader:
        record = {}
        for csv_col, table_col in header_map.items():
            val = row.get(csv_col, "").strip()
            record[table_col] = val if val else None

        if not record.get("canton_code") or not record.get("year"):
            continue

        # Type conversions
        year_val = record.get("year", "")
        try:
            if "-" in str(year_val):
                record["year"] = int(str(year_val).split("-")[0])
            else:
                record["year"] = int(year_val)
        except (ValueError, TypeError):
            continue

        for col in ("total_dwellings", "vacant_dwellings", "bfs_commune_number"):
            if record.get(col) is not None:
                try:
                    record[col] = int(record[col])
                except (ValueError, TypeError):
                    record[col] = None

        if record.get("vacancy_rate_pct") is not None:
            try:
                record["vacancy_rate_pct"] = float(record["vacancy_rate_pct"])
            except (ValueError, TypeError):
                record["vacancy_rate_pct"] = None

        # Add source
        record["source"] = "BFS Leerwohnungszaehlung"
        records.append(record)

    return records


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def main():
    lamap_url = os.environ.get("LAMAP_SUPABASE_URL", "")
    lamap_key = os.environ.get("LAMAP_SUPABASE_SERVICE_KEY", "")
    lamap_schema = os.environ.get("LAMAP_SCHEMA", "bronze")
    camelote_url = os.environ.get("CAMELOTE_SUPABASE_URL", "")
    camelote_key = os.environ.get("CAMELOTE_SUPABASE_KEY", "")

    if not lamap_url or not lamap_key:
        print("ERROR: LAMAP_SUPABASE_URL and LAMAP_SUPABASE_SERVICE_KEY are required")
        sys.exit(1)

    print("=" * 60)
    print("  BFS Vacancy Rates Pipeline")
    print(f"  Target: {lamap_schema}.{TABLE}")
    print("=" * 60)

    # ── Row count BEFORE ──
    rows_before = get_row_count(lamap_url, lamap_key, lamap_schema)
    print(f"\n  Rows before: {rows_before:,}" if rows_before is not None else "\n  Rows before: unknown")

    # ── Download & Parse ──
    start = time.time()
    records = []
    source_url = None

    # Strategy 1: BFS DAM API (primary — XLSX with multi-sheet canton data)
    print("\n  Trying BFS DAM API assets (XLSX)...")
    for asset_id in BFS_ASSET_IDS:
        asset_url = BFS_ASSET_URL_TEMPLATE.format(asset_id)
        try:
            xlsx_bytes = download_xlsx(asset_url)
            print("\n  Parsing XLSX...")
            records = parse_xlsx_to_records(xlsx_bytes)
            if records:
                source_url = asset_url
                print(f"  Success: asset {asset_id} → {len(records):,} records")
                break
            else:
                print(f"  Asset {asset_id}: downloaded but parsed 0 records, trying next...")
        except requests.exceptions.HTTPError:
            print(f"  Asset {asset_id} not available, trying next...")
            continue
        except Exception as e:
            print(f"  Asset {asset_id} parse error: {e}, trying next...")
            continue

    # Strategy 2: opendata.swiss CSV search (fallback)
    if not records:
        print("\n  BFS DAM assets unavailable, searching opendata.swiss for CSV...")
        csv_url = find_csv_url_opendata()
        if csv_url:
            try:
                text = download_csv(csv_url)
                print("\n  Parsing CSV...")
                records = parse_csv_to_records(text)
                if records:
                    source_url = csv_url
            except requests.exceptions.HTTPError as e:
                print(f"  Warning: opendata.swiss URL failed ({e})")

    print(f"  Parsed {len(records):,} records total")

    if not records:
        print("  ERROR: No records parsed from any source")
        sys.exit(1)

    # Normalise keys
    all_keys = set()
    for rec in records:
        all_keys |= rec.keys()
    for rec in records:
        for k in all_keys:
            rec.setdefault(k, None)

    # ── Upsert ──
    print(f"\n  Upserting {len(records):,} records...")
    upserted = batch_upsert(
        url=lamap_url,
        key=lamap_key,
        table=TABLE,
        records=records,
        conflict_column=CONFLICT_COLUMN,
        schema=lamap_schema,
        batch_size=BATCH_SIZE,
    )

    elapsed = time.time() - start

    # ── Row count AFTER ──
    rows_after = get_row_count(lamap_url, lamap_key, lamap_schema)

    # ── Summary ──
    print(f"\n{'=' * 60}")
    print("  IMPORT COMPLETE")
    print(f"  Source:          {source_url}")
    print(f"  Records parsed:  {len(records):,}")
    print(f"  Rows upserted:   {upserted:,}")
    print(f"  Rows before:     {rows_before:,}" if rows_before is not None else "  Rows before:     unknown")
    print(f"  Rows after:      {rows_after:,}" if rows_after is not None else "  Rows after:      unknown")
    if rows_before is not None and rows_after is not None:
        delta = rows_after - rows_before
        print(f"  Net new:         {delta:,}")
    print(f"  Duration:        {elapsed:.1f}s")
    print("=" * 60)

    if upserted == 0:
        print("  FAILED: Zero rows upserted!")
        sys.exit(1)

    update_dataset_meta(
        camelote_url, camelote_key, DATASET_CODE,
        record_count=rows_after,
        status="active",
    )


if __name__ == "__main__":
    main()

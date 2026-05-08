#!/usr/bin/env python3
"""
BFS IMPI — Swiss Residential Property Price Index — Import Pipeline

Source:  BFS DAM, latest "Indexwerte" xlsx asset (auto-discovered each run via the
         DAM list API filtered on inquiry=IMPI & articleType=xlsx).
         The latest publication's index-level time series sheet (T2) covers
         Q1 2017 → current quarter, base 4Q2019 = 100, three property types
         (Wohneigentum total / EFH single-family / EGW apartments) ×
         (Total + GemeindeTyp 1..5) = 18 subindex columns.

Target:  bronze_ch.bfs_impi on re-LLM (one row per period × subindex × base_year)
Conflict key: (period, property_type, commune_type, base_year)
Cadence: quarterly (BFS publishes ~25th of the second month after each quarter)

DATA SAFETY:
    - UPSERT only (INSERT ... ON CONFLICT DO UPDATE).
    - Never truncates or deletes existing data.
    - Row count should only go UP (or stay the same on idempotent re-runs).

Environment variables:
    RE_LLM_SUPABASE_URL              - re-LLM Supabase project URL (required)
    RE_LLM_SUPABASE_SERVICE_ROLE_KEY - service_role key (required)
    RE_LLM_SCHEMA                    - target schema (default: bronze_ch)
"""

from __future__ import annotations

import io
import os
import re
import sys

import openpyxl
import requests

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
from shared.supabase_client import batch_upsert  # noqa: E402

# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

DAM_LIST_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets"
DAM_MASTER_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/{asset_id}/master"

TABLE_NAME = "bfs_impi"

# Map xlsx column position (0-indexed within data block) to (property_type, commune_type)
# T2 sheet header row 10 layout:
#   col 0 = quarter label
#   col 1 = Wohneigentum Total       cols 2-6 = Wohneigentum GemeindeTyp 1..5
#   col 7 = EFH Total                cols 8-12 = EFH GemeindeTyp 1..5
#   col 13 = EGW Total               cols 14-18 = EGW GemeindeTyp 1..5
COLUMN_LAYOUT: list[tuple[int, str, str]] = [
    (1,  "wohneigentum", "total"),
    (2,  "wohneigentum", "communeType1"),
    (3,  "wohneigentum", "communeType2"),
    (4,  "wohneigentum", "communeType3"),
    (5,  "wohneigentum", "communeType4"),
    (6,  "wohneigentum", "communeType5"),
    (7,  "efh",          "total"),
    (8,  "efh",          "communeType1"),
    (9,  "efh",          "communeType2"),
    (10, "efh",          "communeType3"),
    (11, "efh",          "communeType4"),
    (12, "efh",          "communeType5"),
    (13, "egw",          "total"),
    (14, "egw",          "communeType1"),
    (15, "egw",          "communeType2"),
    (16, "egw",          "communeType3"),
    (17, "egw",          "communeType4"),
    (18, "egw",          "communeType5"),
]

# Match labels like "Q1 2026" / "1. Quartal 2026"
_PERIOD_RE = re.compile(r"Q([1-4])\s*(\d{4})", re.IGNORECASE)


# ──────────────────────────────────────────────────────────────
# Source discovery
# ──────────────────────────────────────────────────────────────

def find_latest_indexwerte_asset() -> tuple[int, str, str]:
    """Query DAM API for the most recent IMPI 'Indexwerte' xlsx.

    Returns (asset_id, title, base_year_label).
    """
    r = requests.get(
        DAM_LIST_URL,
        params={"inquiry": "IMPI", "articleType": "xlsx", "size": 50},
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=30,
    )
    r.raise_for_status()
    data = r.json().get("data", [])

    candidates = []
    for a in data:
        am = a.get("bfs", {}).get("articleModel", {}) or {}
        desc = a.get("description", {}) or {}
        titles = desc.get("titles", {}) or {}
        main = str(titles.get("main", ""))
        if "TABL" in str(am.get("code", "")) and main.startswith("Indexwerte"):
            asset_id = a.get("ids", {}).get("damId")
            sub = str(titles.get("sub", ""))
            base = ""
            m = re.search(r"Basis:\s*(Q\d\s*\d{4})", sub)
            if m:
                base = m.group(1).replace(" ", "_")
            candidates.append((asset_id, main, base))

    if not candidates:
        raise RuntimeError(
            "No IMPI 'Indexwerte' xlsx asset found in DAM list — source may have moved."
        )

    # Sort by asset id desc — newer publications get higher dam IDs.
    candidates.sort(key=lambda x: -(x[0] or 0))
    return candidates[0]


# ──────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────

def parse_indexwerte_xlsx(xlsx_bytes: bytes, base_year: str) -> list[dict]:
    """Parse the T2 sheet (or first non-translation sheet) into long-format rows."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)

    data_sheet = next((s for s in wb.sheetnames if s != "Uebersetzungen"), None)
    if data_sheet is None:
        raise RuntimeError("No data sheet found in xlsx (only Uebersetzungen).")

    ws = wb[data_sheet]
    rows = list(ws.iter_rows(values_only=True))

    out: list[dict] = []
    for row in rows[11:]:
        period_label = row[0] if row else None
        if not period_label:
            continue
        m = _PERIOD_RE.search(str(period_label))
        if not m:
            # End of data (e.g. footer row "Auskunft: …")
            continue
        quarter, year = int(m.group(1)), int(m.group(2))
        period = f"{year}-Q{quarter}"

        for col_idx, prop, ct in COLUMN_LAYOUT:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == "":
                continue
            try:
                fv = float(val)
            except (TypeError, ValueError):
                continue
            out.append(
                {
                    "period": period,
                    "year": year,
                    "quarter": quarter,
                    "property_type": prop,
                    "commune_type": ct,
                    "base_year": base_year or "Q4_2019",
                    "index_value": fv,
                }
            )
    return out


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def _required(name: str) -> str:
    v = os.environ.get(name, "")
    if not v:
        print(f"ERROR: env var {name} required")
        sys.exit(1)
    return v


def main():
    url = _required("RE_LLM_SUPABASE_URL")
    key = _required("RE_LLM_SUPABASE_SERVICE_ROLE_KEY")
    schema = os.environ.get("RE_LLM_SCHEMA", "bronze_ch")

    print("=" * 60)
    print("  BFS IMPI Import Pipeline")
    print(f"  Target: {schema}.{TABLE_NAME} on {url}")
    print("=" * 60)

    print("\n[1/3] Discovering latest IMPI Indexwerte asset…")
    asset_id, title, base = find_latest_indexwerte_asset()
    print(f"  asset_id = {asset_id}")
    print(f"  title    = {title}")
    print(f"  base     = {base}")

    print("\n[2/3] Downloading xlsx…")
    r = requests.get(
        DAM_MASTER_URL.format(asset_id=asset_id),
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    r.raise_for_status()
    print(f"  bytes received: {len(r.content):,}")

    rows = parse_indexwerte_xlsx(r.content, base_year=base)
    print(f"  parsed rows: {len(rows):,}")

    if not rows:
        print("ERROR: no rows parsed — refusing to upsert empty payload.")
        sys.exit(1)

    print("\n[3/3] Upserting …")
    n = batch_upsert(
        url=url,
        key=key,
        table=TABLE_NAME,
        records=rows,
        conflict_column="period,property_type,commune_type,base_year",
        schema=schema,
        batch_size=500,
    )
    print(f"  upserted: {n}")

    print("\n" + "=" * 60)
    print("  IMPORT COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()

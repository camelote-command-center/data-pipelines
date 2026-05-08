#!/usr/bin/env python3
"""
BFS Rent CPI subindex (MPI MULTIBASIS) — Import Pipeline

Source:  BFS DAM, latest "IPC, indice des loyers sur toutes les bases d'indice
         [MPI MULTIBASIS]" xlsx asset (auto-discovered each run via the DAM
         list API filtered on inquiry=LIK; multibasis FR variant preferred).

         The xlsx contains a `MPI_m` sheet (monthly values, 1939-08 → current
         month) with 11 index-base columns. Each column holds the rent index
         rebased to a different reference date = 100. Source publishes monthly.

         The yearly averages sheet (MPI_y) and quarterly variation rates sheet
         (%) are derivations of MPI_m and intentionally not ingested — both
         can be derived in SQL from the monthly base-1939 series.

Target:  bronze_ch.bfs_rent_cpi (long-format: one row per period × index_base)
Conflict key: (period_date, base_date)
Cadence: monthly (BFS publishes the CPI early in each month)

Environment variables:
    RE_LLM_SUPABASE_URL              - re-LLM Supabase project URL (required)
    RE_LLM_SUPABASE_SERVICE_ROLE_KEY - service_role key (required)
    RE_LLM_SCHEMA                    - target schema (default: bronze_ch)
"""

from __future__ import annotations

import io
import os
import sys
from datetime import date, datetime

import openpyxl
import requests

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
from shared.supabase_client import batch_upsert  # noqa: E402

# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

DAM_LIST_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets"
DAM_MASTER_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/{asset_id}/master"

TABLE_NAME = "bfs_rent_cpi"


# ──────────────────────────────────────────────────────────────
# Source discovery
# ──────────────────────────────────────────────────────────────

def find_latest_mpi_multibasis_asset() -> tuple[int, str]:
    """Pick the most recent FR MPI MULTIBASIS rent xlsx.

    The asset id is rotated every release. Search by inquiry=LIK and filter
    by title (must contain 'indice des loyers' AND 'MPI MULTIBASIS') and
    language=FR. Highest dam id = newest.
    """
    candidates: list[tuple[int, str]] = []
    for skip in range(0, 200, 20):
        r = requests.get(
            DAM_LIST_URL,
            params={"inquiry": "LIK", "articleType": "xlsx", "size": 20, "skip": skip},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json().get("data", [])
        if not data:
            break
        for a in data:
            am = a.get("bfs", {}).get("articleModel", {}) or {}
            desc = a.get("description", {}) or {}
            titles = desc.get("titles", {}) or {}
            main = str(titles.get("main", ""))
            lang = str(desc.get("language", ""))
            if (am.get("code") == "TABL"
                    and lang == "FR"
                    and "indice des loyers" in main.lower()
                    and "MULTIBASIS" in main):
                asset_id = a.get("ids", {}).get("damId")
                candidates.append((asset_id, main))

    if not candidates:
        raise RuntimeError(
            "No FR MPI MULTIBASIS rent xlsx found in DAM (searched first 200 LIK assets)."
        )

    candidates.sort(key=lambda x: -(x[0] or 0))
    return candidates[0]


# ──────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────

def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def parse_xlsx(xlsx_bytes: bytes) -> list[dict]:
    """Parse the MPI_m sheet into long format.

    Layout:
        r3, col 0       = 'Datum / Date'
        r3, cols 1..11  = base reference dates (YYYY-MM-01) for each index base
        r4..end:
            col 0       = period date (YYYY-MM-01)
            cols 1..11  = index value rebased to col-N base = 100
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    if "MPI_m" not in wb.sheetnames:
        raise RuntimeError(f"MPI_m sheet not found; sheets={wb.sheetnames}")

    ws = wb["MPI_m"]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 5:
        raise RuntimeError("MPI_m sheet has too few rows.")

    # Header row 3: extract base dates from cols 1..11
    header = rows[3]
    base_cols: list[tuple[int, date]] = []
    for col_idx in range(1, len(header)):
        bd = _to_date(header[col_idx])
        if bd is not None:
            base_cols.append((col_idx, bd))

    if not base_cols:
        raise RuntimeError("No base date columns found in header row 3.")

    out: list[dict] = []
    for r_idx in range(4, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        period_date = _to_date(row[0])
        if period_date is None:
            continue
        period = f"{period_date.year}-{period_date.month:02d}"

        for col_idx, base_date in base_cols:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == "":
                continue
            try:
                fv = float(val)
            except (TypeError, ValueError):
                continue
            out.append({
                "period": period,
                "period_date": period_date.isoformat(),
                "base_date": base_date.isoformat(),
                "base_label": f"{base_date.isoformat()} = 100",
                "index_value": fv,
            })

    return out


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def _required(name: str) -> str:
    v = os.environ.get(name, "")
    if not v:
        print(f"ERROR: env var {name} required")
        sys.exit(1)
    return v


def main():
    url = _required("RE_LLM_SUPABASE_URL")
    key = _required("RE_LLM_SUPABASE_SERVICE_ROLE_KEY")
    schema = os.environ.get("RE_LLM_SCHEMA", "bronze_ch")

    print("=" * 60)
    print("  BFS Rent CPI (MPI MULTIBASIS) Import Pipeline")
    print(f"  Target: {schema}.{TABLE_NAME} on {url}")
    print("=" * 60)

    print("\n[1/3] Discovering latest MPI MULTIBASIS asset…")
    asset_id, title = find_latest_mpi_multibasis_asset()
    print(f"  asset_id = {asset_id}")
    print(f"  title    = {title}")

    print("\n[2/3] Downloading xlsx…")
    r = requests.get(
        DAM_MASTER_URL.format(asset_id=asset_id),
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    r.raise_for_status()
    print(f"  bytes received: {len(r.content):,}")

    rows = parse_xlsx(r.content)
    print(f"  parsed rows: {len(rows):,}")
    if rows:
        bases = sorted(set(r["base_date"] for r in rows))
        periods = sorted(set(r["period"] for r in rows))
        print(f"  bases: {len(bases)} ({bases[0]} … {bases[-1]})")
        print(f"  period range: {periods[0]} … {periods[-1]} ({len(periods)} months)")

    if not rows:
        print("ERROR: no rows parsed — refusing to upsert empty payload.")
        sys.exit(1)

    print("\n[3/3] Upserting …")
    n = batch_upsert(
        url=url,
        key=key,
        table=TABLE_NAME,
        records=rows,
        conflict_column="period_date,base_date",
        schema=schema,
        batch_size=1000,
    )
    print(f"  upserted: {n}")

    print("\n" + "=" * 60)
    print("  IMPORT COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
